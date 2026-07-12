"""Gera as planilhas auditaveis de 08_REFERENCIAS/ a partir das fontes de verdade.

Uso: python scripts/python/13_gerar_planilha_referencias.py

Fontes:
    latex-artigo/references.bib                          -> as 35 referencias citadas no artigo
    07_SINTESE_TEMATICA/matriz_base_nucleo_final_104.csv  -> os 104 registros do nucleo final

Saidas (em 08_REFERENCIAS/): copia do .bib, e CSV/XLSX de cada fonte acima.
Ver 08_REFERENCIAS/README.md para a relacao entre as duas listas.
"""
import csv
import re
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
BIB = ROOT / "latex-artigo" / "references.bib"
NUCLEO_104 = ROOT / "07_SINTESE_TEMATICA" / "matriz_base_nucleo_final_104.csv"
DEST = ROOT / "08_REFERENCIAS"


def parse_bib(path):
    texto = path.read_text(encoding="utf-8")
    entradas = []
    for m in re.finditer(r"@(\w+)\{([^,\n]+),", texto):
        tipo, chave = m.group(1), m.group(2).strip()
        depth, i = 1, m.end()
        while depth > 0 and i < len(texto):
            if texto[i] == "{":
                depth += 1
            elif texto[i] == "}":
                depth -= 1
            i += 1
        corpo = texto[m.end():i - 1]
        campos = {}
        for fm in re.finditer(r"(\w+)\s*=\s*\{((?:[^{}]|\{[^{}]*\})*)\}", corpo):
            campos[fm.group(1).strip().lower()] = re.sub(r"\s+", " ", fm.group(2)).strip()
        entradas.append({"chave": chave, "tipo": tipo, **campos})
    return entradas


def formatar_autores(autores):
    if not autores:
        return ""
    return "; ".join(a.strip().strip("{}") for a in autores.split(" and "))


def gerar_csv_referencias_citadas(destino_csv):
    entradas = parse_bib(BIB)
    colunas = ["chave", "autores", "ano", "titulo", "fonte", "volume", "numero",
               "paginas", "tipo", "doi", "url", "instituicao"]
    with open(destino_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(colunas)
        for e in sorted(entradas, key=lambda x: x.get("author", x.get("chave", ""))):
            fonte = e.get("journal") or e.get("booktitle") or e.get("institution") or ""
            doi = e.get("doi", "")
            w.writerow([
                e.get("chave", ""), formatar_autores(e.get("author", "")), e.get("year", ""),
                e.get("title", ""), fonte, e.get("volume", ""), e.get("number", ""),
                e.get("pages", ""), e.get("tipo", ""), doi,
                ("https://doi.org/" + doi) if doi else e.get("url", ""), e.get("institution", ""),
            ])
    return len(entradas)


def csv_para_xlsx(origem_csv, destino_xlsx, delimitador, nome_aba):
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba[:31]
    with open(origem_csv, encoding="utf-8-sig", newline="") as f:
        for i, row in enumerate(csv.reader(f, delimiter=delimitador), start=1):
            ws.append(row)
            if i == 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        largura = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(60, max(10, largura + 2))
    wb.save(destino_xlsx)


if __name__ == "__main__":
    DEST.mkdir(exist_ok=True)

    shutil.copy(BIB, DEST / "referencias_citadas_artigo.bib")

    csv_citadas = DEST / "referencias_citadas_artigo.csv"
    n = gerar_csv_referencias_citadas(csv_citadas)
    csv_para_xlsx(csv_citadas, DEST / "referencias_citadas_artigo.xlsx", ";", "Referencias citadas")
    print(f"Referencias citadas: {n} entradas")

    shutil.copy(NUCLEO_104, DEST / "nucleo_final_104_registros.csv")
    csv_para_xlsx(DEST / "nucleo_final_104_registros.csv", DEST / "nucleo_final_104_registros.xlsx",
                  ",", "Nucleo final 104")
    print("Nucleo final 104 registros: copiado e convertido")
