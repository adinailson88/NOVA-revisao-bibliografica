"""Gera planilhas XLSX de referencias e dos nucleos tematicos para auditoria externa.

Uso: python scripts/python/14_gerar_planilhas_auditoria_referencias.py
(rodar depois de scripts/python/12_gerar_lista_referencias.py)

Gera em referencias/:
    lista_referencias.xlsx          -> copia em XLSX de lista_referencias.csv
    nucleo_final_121_registros.csv  -> copia do nucleo tematico vigente
    nucleo_final_121_registros.xlsx -> versao XLSX do nucleo vigente
    nucleo_final_104_registros.csv  -> copia historica do nucleo original
    nucleo_final_104_registros.xlsx -> versao XLSX do nucleo historico

O nucleo vigente de 121 registros fundamenta a sintese tematica atual. O nucleo de 104
permanece preservado para rastreabilidade e comparacao, sem ser apresentado como produto
vigente.
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
REFERENCIAS = ROOT / "referencias"
NUCLEO_121 = (
    ROOT
    / "latex-artigo"
    / "fontes"
    / "nucleo_final_pos_auditoria_resumos_v2_sensibilidade.csv"
)
NUCLEO_104 = ROOT / "07_SINTESE_TEMATICA" / "matriz_base_nucleo_final_104.csv"


def contar_registros(origem_csv: Path) -> int:
    with origem_csv.open(encoding="utf-8-sig", newline="") as arquivo:
        linhas = list(csv.DictReader(arquivo))
    ids = [linha["id_unico"] for linha in linhas]
    if len(ids) != len(set(ids)):
        raise SystemExit(f"id_unico duplicado em {origem_csv}")
    return len(linhas)


def csv_para_xlsx(origem_csv, destino_xlsx, delimitador, nome_aba):
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba[:31]
    with open(origem_csv, encoding="utf-8-sig", newline="") as f:
        for i, row in enumerate(csv.reader(f, delimiter=delimitador), start=1):
            ws.append(row)
            if i == 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        largura = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            60, max(10, largura + 2)
        )
    wb.save(destino_xlsx)


def publicar_nucleo(origem: Path, total_esperado: int, rotulo: str) -> None:
    total = contar_registros(origem)
    if total != total_esperado:
        raise SystemExit(
            f"Total divergente no nucleo {rotulo}: esperado {total_esperado}, encontrado {total}."
        )
    destino_csv = REFERENCIAS / f"nucleo_final_{total_esperado}_registros.csv"
    destino_csv.write_bytes(origem.read_bytes())
    csv_para_xlsx(
        destino_csv,
        REFERENCIAS / f"nucleo_final_{total_esperado}_registros.xlsx",
        ",",
        f"Nucleo {rotulo} {total_esperado}",
    )
    print(f"nucleo_final_{total_esperado}_registros.csv/.xlsx gerados")


if __name__ == "__main__":
    lista_csv = REFERENCIAS / "lista_referencias.csv"
    if not lista_csv.exists():
        raise SystemExit(
            "referencias/lista_referencias.csv nao encontrado. "
            "Rode antes: python scripts/python/12_gerar_lista_referencias.py"
        )
    csv_para_xlsx(
        lista_csv,
        REFERENCIAS / "lista_referencias.xlsx",
        ",",
        "Referencias citadas",
    )
    print("lista_referencias.xlsx gerado")

    publicar_nucleo(NUCLEO_121, 121, "vigente")
    publicar_nucleo(NUCLEO_104, 104, "historico")
